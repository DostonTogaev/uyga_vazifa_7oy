import csv

import writer
from django.conf import settings
from django.contrib import messages
from django.core.mail import send_mail
from openpyxl import Workbook
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models.functions import Cast
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.db.models import Q, TextField

from app import forms
from customer.forms import CustomerModelForm, ShareMail
from customer.models import Customer
import json
from django.views.generic import TemplateView
from django.views import View

# Create your views here.

class CustomerTemplateView(TemplateView):
    template_name = 'customer/customer-list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customers = Customer.objects.all()
        context['customers'] = customers
        return context

class CustomerListView(View):
    def get(self, request):
        search_query = request.GET.get('search')
        if search_query:
            customer_list = Customer.objects.filter(
                Q(full_name__icontains=search_query) | Q(address__icontains=search_query))
        else:
            customer_list = Customer.objects.all()

        page = request.GET.get('page', '')
        paginator = Paginator(customer_list, 3)
        try:
            customer_list = paginator.page(page)
        except PageNotAnInteger:
            customer_list = paginator.page(1)
        except EmptyPage:
            customer_list = paginator.page(paginator.num_pages)
        context = {
            'customer_list': customer_list,

        }
        return render(request, 'customer/customer-list.html', context)


'''def customers(request):
    search_query = request.GET.get('search')
    if search_query:
        customer_list = Customer.objects.filter(
            Q(full_name__icontains=search_query) | Q(address__icontains=search_query))
    else:
        customer_list = Customer.objects.all()

    page = request.GET.get('page', '')
    paginator = Paginator(customer_list, 3)
    try:
        customer_list = paginator.page(page)
    except PageNotAnInteger:
        customer_list = paginator.page(1)
    except EmptyPage:
        customer_list = paginator.page(paginator.num_pages)
    context = {
        'customer_list': customer_list,

    }
    return render(request, 'customer/customer-list.html', context)'''

class CustomerDetailView(View):
    def get(self, request, customer_id):
        customer = Customer.objects.get(id=customer_id)
        context = {
            'customer': customer
        }
        return render(request, 'customer/customer-detail.html', context)
'''def customer_detail(request, customer_id):
    customer = Customer.objects.get(id=customer_id)
    context = {
        'customer': customer
    }
    return render(request, 'customer/customer-detail.html', context)'''

class CustomerAddTemplateView(TemplateView):
    template_name = 'customer/add-customer.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = Customer.objects.get(id=self.kwargs['customer_id'])
        context['customer'] = CustomerModelForm(instance=customer)
        return context

    def post(self, request, *args, **kwargs):
        context = self.get_context_data()

class CustomerAddView(View):
    def get(self, request):
        form = CustomerModelForm()
        return render(request, 'customer/add-customer.html', {'form': form})
    def post(self, request):
        form = CustomerModelForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customers')


'''def add_customer(request):
    form = CustomerModelForm()
    if request.method == 'POST':
        form = CustomerModelForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('customers')

    context = {
        'form': form,
    }

    return render(request, 'customer/add-customer.html', context)'''

class CustomerDeleteView(View):
    def get(self, request, pk):
        customer = Customer.objects.get(id=pk)
        if customer:
            customer.delete()
            return redirect('customers')

'''def delete_customer(request, pk):
    customer = Customer.objects.get(id=pk)
    if customer:
        customer.delete()
        messages.add_message(
            request,
            messages.SUCCESS,
            'Customer successfully deleted'
        )
        return redirect('customers')'''

class CustomerEditView(View):
    def get(self, request, pk):
        customer = Customer.objects.get(id=pk)
        form = CustomerModelForm(instance=customer)
        return render(request, 'customer/update-customer.html', {'form': form})

    def post(self, request, pk):
        customer = Customer.objects.get(id=pk)
        form = CustomerModelForm(request.POST, instance=customer)

        if form.is_valid():
            form.save()
            return redirect('customer_detail', pk)
'''def edit_customer(request, pk):
    customer = Customer.objects.get(id=pk)
    form = CustomerModelForm(instance=customer)
    if request.method == 'POST':
        form = CustomerModelForm(instance=customer, data=request.POST, files=request.FILES)
        if form.is_valid():
            form.save()

            return redirect('customers')
    context = {
        'form': form,
    }
    return render(request, 'customer/update-customer.html', context)'''

def export_data(request):
    format = request.GET.get('format', 'csv')
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=customers.csv'
        writer = csv.writer(response)
        writer.writerow(['Id', 'Full Name', 'Email', 'Phone Number', 'Address'])
        for customer in Customer.objects.all():
            writer.writerow([customer.id, customer.full_name, customer.email, customer.phone_number, customer.address])


    elif format == 'json':
        response = HttpResponse(content_type='application/json')
        data = list(Customer.objects.all().values('full_name', 'email', 'phone_number', 'address'))
        # response.content = json.dumps(data, indent=4)
        response.write(json.dumps(data, indent=4))
        response['Content-Disposition'] = 'attachment; filename=customers.json'
    elif format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Customers'

        # Write header row
        header = ['ID', 'Name', 'Email', 'Phone Number', 'Address']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        queryset = Customer.objects.all().values_list('id', 'full_name', 'email', 'phone_number', 'address')
        for row_num, row in enumerate(queryset, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)
    else:
        response = HttpResponse(status=404)
        response.content = 'Bad request'

    return response

def share_mail(request):
    sent = False
    if request.method == 'Post':
        form = ShareMail(request.POST)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            body = form.cleaned_data['body']
            form_email = settings.DEFAULT_FROM_EMAIL
            recipients = form.cleaned_data['recipients']
            send_mail(subject, body, form_email, recipients, fail_silently=False)
            sent = True

    else:
        form = ShareMail()

    return render(request, 'email/share_email.html', {'form': form, 'sent': sent})
